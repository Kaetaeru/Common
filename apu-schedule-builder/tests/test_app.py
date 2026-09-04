import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import State, build_normalized, conflict, generate_schedules, schedule_metrics, section_slots, state_score


class APUScheduleBuilderTests(unittest.TestCase):
    def make_files(self, root: Path):
        timetable = root / "timetable.xlsx"
        subjects = root / "subjects.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"
        ws.append(["AY2026 Fall Timetable"])
        ws.append(["Term", "Day of Week", "Period", "Classroom", "In-Person or Online", "Course code", "Subject CD", "Subject Name", "Instructor", "Lang.", "Semester", "Field", "Area of Study"])
        ws.append(["Semester", "Mon", 3, "J101", "In-Person", 1001, "038001EA", "Semester Course", "Prof A", "E", 1, "Major", "AF"])
        ws["H3"].hyperlink = "https://syllabus.apu.ac.jp/syllabus/s/a-syllabus/a0ZQTEST000000001/20261001?language=en_US"
        # Continuation row intentionally leaves merged-style metadata blank.
        ws.append([None, "Thu", 3, "J101", None, None, None, None, None, None, None, None, None])
        ws.append(["1Q", "Mon", 3, "J102", "In-Person", 1002, "038002EA", "Q1 Course", "Prof B", "E", 1, "Major", "AF"])
        ws.append(["2Q", "Mon", 3, "J103", "In-Person", 1003, "038003EA", "Q2 Course", "Prof C", "E", 1, "Major", "AF"])
        ws.append(["Semester", "Tue", 2, "J104", "In-Person", 1004, "038004EA", "Preferred Course", "Prof D", "E", 3, "Major", "M"])
        ws.append(["Semester", "Fri", 2, "J105", "In-Person", 1005, "038005EA", "Late Eligibility", "Prof E", "E", 7, "Major", "M"])
        wb.save(timetable)

        wb = Workbook()
        ws = wb.active
        ws.title = "Subject List"
        ws.append(["Notes"])
        ws.append(["Subject Code", "Subject Name", "Credits", "Semester", "Field", "Area of Study", "Prerequisite Subjects for E Students"])
        ws.append(["038001", "Semester Course", 2, 1, "Major", "AF", ""])
        ws.append(["038002", "Q1 Course", 2, 1, "Major", "AF", ""])
        ws.append(["038003", "Q2 Course", 2, 1, "Major", "AF", ""])
        ws.append(["038004", "Preferred Course", 2, 3, "Major", "M", "Semester Course"])
        ws.append(["038005", "Late Eligibility", 2, 7, "Major", "M", ""])
        wb.save(subjects)
        return timetable, subjects

    def test_parser_groups_multiple_meetings_and_joins_credits(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            data = build_normalized("APM", t, s)
            sem = next(x for x in data["sections"] if x["classCode"] == "1001")
            self.assertEqual(len(sem["meetings"]), 2)
            self.assertEqual(sem["credits"], 2)
            self.assertFalse(sem["creditsEstimated"])
            self.assertEqual(sem["syllabusUrl"], "https://syllabus.apu.ac.jp/syllabus/s/a-syllabus/a0ZQTEST000000001/20261001?language=en_US")
            self.assertEqual(data["stats"]["subjects"], 5)
            self.assertEqual(data["term"], "AY2026 Fall")
            self.assertEqual(data["academicYear"], 2026)
            self.assertEqual(data["academicSeason"], "FALL")
            self.assertTrue(data["academicTermDetected"])


    def test_academic_term_is_detected_from_timetable_not_student_semester(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            from openpyxl import load_workbook
            wb = load_workbook(t)
            ws = wb["Timetable"]
            ws["A1"] = "AY2027 Spring Timetable"
            wb.save(t)

            data = build_normalized("APM", t, s)
            self.assertEqual(data["term"], "AY2027 Spring")
            self.assertEqual(data["academicYear"], 2027)
            self.assertEqual(data["academicSeason"], "SPRING")
            self.assertTrue(data["academicTermDetected"])

    def test_known_psychology_syllabus_link_is_attached_by_class_code_and_year(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            from openpyxl import load_workbook
            wb = load_workbook(t)
            ws = wb["Timetable"]
            ws.append(["Semester", "Wed", 2, "J210", "In-Person", 10121, "027019EA", "Psychology", "Prof P", "E", 1, "Liberal", "GCF"])
            wb.save(t)

            wb = load_workbook(s)
            ws = wb["Subject List"]
            ws.append(["027019", "Psychology", 2, 1, "Liberal", "GCF", ""])
            wb.save(s)

            data = build_normalized("APM", t, s)
            psychology = next(x for x in data["sections"] if x["classCode"] == "10121")
            self.assertEqual(psychology["syllabusUrl"], "https://syllabus.apu.ac.jp/syllabus/s/a-syllabus/a0ZQ8000004S4L9MAK/202610121?language=en_US")

    def test_subject_keeps_multiple_class_options_and_each_class_keeps_its_meetings(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            from openpyxl import load_workbook
            wb = load_workbook(t)
            ws = wb["Timetable"]
            ws.append(["Semester", "Tue", 4, "J201", "In-Person", 2001, "038100JA", "Japanese Course", "Prof J1", "J", 1, "Language", ""])
            ws.append([None, "Thu", 4, "J201", None, None, None, None, None, None, None, None, None])
            ws.append(["Semester", "Tue", 5, "J202", "In-Person", 2002, "038100JB", "Japanese Course", "Prof J2", "J", 1, "Language", ""])
            ws.append([None, "Thu", 5, "J202", None, None, None, None, None, None, None, None, None])
            wb.save(t)

            wb = load_workbook(s)
            ws = wb["Subject List"]
            ws.append(["038100", "Japanese Course", 4, 1, "Language", "", ""])
            wb.save(s)

            data = build_normalized("APM", t, s)
            subject = next(x for x in data["subjects"] if x["subjectCode"] == "038100")
            self.assertEqual({x["classCode"] for x in subject["sections"]}, {"2001", "2002"})
            self.assertEqual({len(x["meetings"]) for x in subject["sections"]}, {2})

    def test_quarter_conflict_rules(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            data = build_normalized("APM", t, s)
            by_code = {x["classCode"]: x for x in data["sections"]}
            self.assertTrue(conflict(by_code["1001"], by_code["1002"]))
            self.assertTrue(conflict(by_code["1001"], by_code["1003"]))
            self.assertFalse(conflict(by_code["1002"], by_code["1003"]))
            self.assertEqual(len(section_slots(by_code["1001"])), 4)

    def test_solver_allows_q1_and_q2_same_slot(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            data = build_normalized("APM", t, s)
            result = generate_schedules(data, {
                "semesterLevel": 5,
                "targetCredits": 4,
                "maxCredits": 20,
                "statuses": {"038002": "MUST", "038003": "MUST"},
                "fixedClassCodes": [],
                "autofill": False,
                "earliestPeriod": 1,
                "maxCampusDays": 5,
                "daysOff": [],
            })
            self.assertFalse(result["errors"])
            self.assertTrue(result["results"])
            codes = {c["classCode"] for c in result["results"][0]["courses"]}
            self.assertEqual(codes, {"1002", "1003"})

    def test_solver_rejects_semester_and_q1_same_slot(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            data = build_normalized("APM", t, s)
            result = generate_schedules(data, {
                "semesterLevel": 5,
                "targetCredits": 4,
                "maxCredits": 20,
                "statuses": {"038001": "MUST", "038002": "MUST"},
                "fixedClassCodes": [],
                "autofill": False,
                "earliestPeriod": 1,
                "maxCampusDays": 5,
                "daysOff": [],
            })
            self.assertFalse(result["results"])
            self.assertTrue(result["errors"])

    def test_semester_eligibility_blocks_must_course(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            data = build_normalized("APM", t, s)
            result = generate_schedules(data, {
                "semesterLevel": 5,
                "targetCredits": 2,
                "maxCredits": 20,
                "statuses": {"038005": "MUST"},
                "fixedClassCodes": [],
                "autofill": False,
                "earliestPeriod": 1,
                "maxCampusDays": 5,
                "daysOff": [],
            })
            self.assertFalse(result["results"])
            self.assertIn("no eligible", result["errors"][0])

    def test_blocked_q1_slot_rejects_q1_but_allows_q2_same_time(self):
        with tempfile.TemporaryDirectory() as tmp:
            tfile, sfile = self.make_files(Path(tmp))
            data = build_normalized("APM", tfile, sfile)
            blocked = ["Q1:MON:3"]
            rejected = generate_schedules(data, {
                "semesterLevel": 5, "targetCredits": 2, "maxCredits": 20,
                "statuses": {"038002": "MUST"}, "fixedClassCodes": [],
                "blockedSlots": blocked, "autofill": False,
            })
            self.assertFalse(rejected["results"])
            self.assertIn("disabled time", rejected["errors"][0])

            allowed = generate_schedules(data, {
                "semesterLevel": 5, "targetCredits": 2, "maxCredits": 20,
                "statuses": {"038003": "MUST"}, "fixedClassCodes": [],
                "blockedSlots": blocked, "autofill": False,
            })
            self.assertTrue(allowed["results"])
            self.assertFalse(allowed["errors"])

    def test_blocked_q1_slot_rejects_semester_course(self):
        with tempfile.TemporaryDirectory() as tmp:
            tfile, sfile = self.make_files(Path(tmp))
            data = build_normalized("APM", tfile, sfile)
            result = generate_schedules(data, {
                "semesterLevel": 5, "targetCredits": 2, "maxCredits": 20,
                "statuses": {"038001": "MUST"}, "fixedClassCodes": [],
                "blockedSlots": ["Q1:MON:3"], "autofill": False,
            })
            self.assertFalse(result["results"])
            self.assertIn("disabled time", result["errors"][0])

    def test_fixed_class_respects_semester_eligibility(self):
        with tempfile.TemporaryDirectory() as tmp:
            tfile, sfile = self.make_files(Path(tmp))
            data = build_normalized("APM", tfile, sfile)
            result = generate_schedules(data, {
                "semesterLevel": 5, "targetCredits": 2, "maxCredits": 20,
                "statuses": {}, "fixedClassCodes": ["1005"],
                "blockedSlots": [], "autofill": False,
            })
            self.assertFalse(result["results"])
            self.assertIn("requires semester 7", result["errors"][0])

    def test_fixed_course_reports_disabled_time_conflict(self):
        with tempfile.TemporaryDirectory() as tmp:
            tfile, sfile = self.make_files(Path(tmp))
            data = build_normalized("APM", tfile, sfile)
            result = generate_schedules(data, {
                "semesterLevel": 5, "targetCredits": 2, "maxCredits": 20,
                "statuses": {}, "fixedClassCodes": ["1002"],
                "blockedSlots": ["Q1:MON:3"], "autofill": False,
            })
            self.assertFalse(result["results"])
            self.assertIn("disabled time slot", result["errors"][0])

    def test_schedule_metrics_tracks_longest_daily_gap(self):
        state = State(occupied={("Q1", "MON", 1), ("Q1", "MON", 4), ("Q1", "TUE", 2), ("Q1", "TUE", 3)})
        metrics = schedule_metrics(state)
        self.assertEqual(metrics["gaps"], 2)
        self.assertEqual(metrics["maxGap"], 2)

    def test_language_preference_affects_soft_score(self):
        english = State(chosen=[{"language": "E"}])
        japanese = State(chosen=[{"language": "J"}])
        config = {"targetCredits": 0, "preferredLanguages": ["E"], "earliestPeriod": 1, "latestPeriod": 6, "maxCampusDays": 5, "maxGap": 5}
        self.assertGreater(state_score(english, config, "balanced"), state_score(japanese, config, "balanced"))
    def test_parsed_workbooks_are_released_for_the_refresh_flow(self):
        with tempfile.TemporaryDirectory() as tmp:
            t, s = self.make_files(Path(tmp))
            build_normalized("APM", t, s)
            # Refresh deletes both source files before downloading them again, which
            # Windows refuses while openpyxl still holds the workbook open.
            t.unlink()
            s.unlink()
            self.assertFalse(t.exists())
            self.assertFalse(s.exists())


if __name__ == "__main__":
    unittest.main()
