from __future__ import annotations

import json
import math
import re
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from openpyxl import load_workbook

from syllabus_mapping import load_verified_mapping, parse_direct_syllabus_url

ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
DATA_DIR = ROOT / "data"
SOURCE_DIR = DATA_DIR / "source"
NORMALIZED_DIR = DATA_DIR / "normalized"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)
NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)

DATA_VERSION = "AY2026 Fall / timetable updated 2026-09-01 / subject list updated 2026-03-13"
NORMALIZED_SCHEMA_VERSION = 2

OFFICIAL_FILES = {
    "APS": {
        "timetable": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023APS_Curriculum_26Fall_260901.xlsx",
        "subjects": "https://en.apu.ac.jp/academic/assets/file/aps/subject_list/APS_2023_Subject_List_260313.xlsx",
    },
    "APM": {
        "timetable": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023APM_Curriculum_26Fall_260901.xlsx",
        "subjects": "https://en.apu.ac.jp/academic/assets/file/apm/subject_list/APM_2023_Subject_List_260313.xlsx",
    },
    "ST": {
        "timetable": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023ST_Curriculum_26Fall_260901.xlsx",
        "subjects": "https://en.apu.ac.jp/academic/assets/file/st/subject_list/ST_2023_Subject_List_260313.xlsx",
    },
}

DAY_MAP = {
    "mon": "MON", "monday": "MON", "月": "MON", "月曜": "MON", "月曜日": "MON",
    "tue": "TUE", "tues": "TUE", "tuesday": "TUE", "火": "TUE", "火曜": "TUE", "火曜日": "TUE",
    "wed": "WED", "wednesday": "WED", "水": "WED", "水曜": "WED", "水曜日": "WED",
    "thu": "THU", "thur": "THU", "thurs": "THU", "thursday": "THU", "木": "THU", "木曜": "THU", "木曜日": "THU",
    "fri": "FRI", "friday": "FRI", "金": "FRI", "金曜": "FRI", "金曜日": "FRI",
}
DAYS = ["MON", "TUE", "WED", "THU", "FRI"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def key_text(value: Any) -> str:
    s = clean_text(value).lower()
    return re.sub(r"[\s\n\r\t._/()（）【】\[\]・:：-]+", "", s)


def code_text(value: Any) -> str:
    s = clean_text(value)
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def base_subject_code(value: Any) -> str:
    s = code_text(value).upper()
    match = re.search(r"\d{5,8}", s)
    if match:
        return match.group(0)
    return re.sub(r"[^A-Z0-9]", "", s)


def is_direct_syllabus_url(url: Any, class_code: Any | None = None, academic_year: int | None = None) -> bool:
    text = clean_text(url)
    if not text:
        return False
    try:
        parsed = urlparse(text)
    except ValueError:
        return False
    if parsed.scheme not in {"http", "https"} or parsed.netloc.lower() != "syllabus.apu.ac.jp":
        return False
    parts = [part for part in parsed.path.split("/") if part]
    if "a-syllabus" not in parts or len(parts) < 2:
        return False
    if class_code is not None:
        code = code_text(class_code)
        tail = parts[-1]
        if academic_year is not None:
            if tail != f"{academic_year}{code}":
                return False
        elif not tail.endswith(code):
            return False
    return True


def syllabus_link_from_cells(cells: Iterable[Any], class_code: Any) -> str:
    for cell in cells:
        link = getattr(cell, "hyperlink", None)
        target = getattr(link, "target", None) if link else None
        if is_direct_syllabus_url(target, class_code):
            return clean_text(target)
    return ""


def load_syllabus_link_overrides() -> dict[str, str]:
    """Verified "YYYY:ClassCode" -> direct URL mapping from every repository source."""
    return load_verified_mapping(DATA_DIR)


def apply_syllabus_links(
    sections: Iterable[dict[str, Any]],
    academic_year: int | None,
    overrides: dict[str, str] | None = None,
) -> None:
    if overrides is None:
        overrides = load_syllabus_link_overrides()
    for section in sections:
        class_code = code_text(section.get("classCode"))
        direct = clean_text(section.get("syllabusUrl"))
        if direct and is_direct_syllabus_url(direct, class_code, academic_year):
            continue
        section.pop("syllabusUrl", None)
        if academic_year is None:
            continue
        override = overrides.get(f"{academic_year}:{class_code}", "")
        parsed = parse_direct_syllabus_url(override)
        # The mapping reader already validated the key. For verified grouped
        # aliases, only the URL year must equal the active academic year.
        if parsed and parsed[0] == int(academic_year):
            section["syllabusUrl"] = override


def attach_syllabus_links(data: dict[str, Any]) -> None:
    """Refresh syllabus URLs across a normalized payload with a single mapping read."""
    year = data.get("academicYear")
    overrides = load_syllabus_link_overrides()
    apply_syllabus_links(data.get("sections", []), year, overrides)
    for subject in data.get("subjects", []):
        apply_syllabus_links(subject.get("sections", []), year, overrides)


def detect_academic_term(path: Path) -> dict[str, Any]:
    """Detect AY + Spring/Fall from the timetable workbook itself."""
    candidates: list[str] = [path.name]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            candidates.append(ws.title)
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
                for cell in row:
                    text = clean_text(cell)
                    if text:
                        candidates.append(text)
    finally:
        wb.close()

    joined = " ".join(candidates)
    match = re.search(r"(?:AY\s*)?(20\d{2})\s*(Spring|Fall)", joined, re.IGNORECASE)
    if not match:
        match = re.search(r"(?<!\d)(\d{2})\s*(Spring|Fall)(?![A-Za-z])", joined, re.IGNORECASE)
        if match:
            year = 2000 + int(match.group(1))
            season = match.group(2).title()
            return {"label": f"AY{year} {season}", "year": year, "season": season.upper(), "detected": True}
        return {"label": "Academic term unknown", "year": None, "season": None, "detected": False}

    year = int(match.group(1))
    season = match.group(2).title()
    return {"label": f"AY{year} {season}", "year": year, "season": season.upper(), "detected": True}


def normalize_term(value: Any) -> str:
    s = key_text(value)
    if not s:
        return "SEMESTER"
    if "1q" in s or "quarter1" in s or "firstquarter" in s or "第1" in s:
        return "Q1"
    if "2q" in s or "quarter2" in s or "secondquarter" in s or "第2" in s:
        return "Q2"
    if "semester" in s or "セメスター" in s:
        return "SEMESTER"
    if "session" in s or "summer" in s or "winter" in s:
        return "SESSION"
    return clean_text(value).upper()


def normalize_day(value: Any) -> str | None:
    s = key_text(value)
    if s in DAY_MAP:
        return DAY_MAP[s]
    for key, day in DAY_MAP.items():
        if s.startswith(key):
            return day
    return None


def parse_period(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        p = int(value)
        return p if 1 <= p <= 6 else None
    match = re.search(r"[1-6]", clean_text(value))
    return int(match.group(0)) if match else None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", clean_text(value))
    return float(match.group(0)) if match else None


def parse_semester_min(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        return n if 1 <= n <= 8 else None
    s = clean_text(value)
    match = re.search(r"([1-8])", s)
    return int(match.group(1)) if match else None


def normalize_mode(value: Any) -> str:
    s = key_text(value)
    if "ondemand" in s or "オンデマンド" in s:
        return "ON_DEMAND"
    if "online" in s or "zoom" in s or "オンライン" in s:
        return "ONLINE"
    return "IN_PERSON" if s else "UNKNOWN"


HEADER_ALIASES = {
    "term": ["term", "開講期間", "授業期間"],
    "day": ["dayofweek", "day", "曜日"],
    "period": ["period", "時限"],
    "classroom": ["classroom", "教室"],
    "mode": ["inpersonoronline", "classformat", "授業形態", "対面オンライン"],
    "class_code": ["coursecode", "classcode", "講義コード", "授業コード"],
    "subject_cd": ["subjectcd", "subjectcode", "科目コード"],
    "name": ["subjectname", "coursename", "科目名"],
    "instructor": ["instructor", "teacher", "教員", "担当教員"],
    "language": ["lang", "language", "言語"],
    "semester": ["semester", "セメスター"],
    "field": ["field", "分野", "科目分野"],
    "area": ["areaofstudy", "area", "学修分野", "専門分野"],
    "credits": ["credits", "credit", "単位"],
}

SUBJECT_HEADER_ALIASES = {
    "subject_code": ["subjectcode", "subjectcd", "科目コード"],
    "name": ["subjectname", "coursename", "科目名"],
    "credits": ["credits", "credit", "単位"],
    "semester": ["semester", "セメスター"],
    "field": ["field", "分野", "科目分野"],
    "area": ["areaofstudy", "area", "学修分野", "専門分野"],
    "prereq_jst": ["prerequisitesubjectsforjststudents", "jststudents", "jst"],
    "prereq_jat": ["prerequisitesubjectsforjatstudents", "jatstudents", "jat"],
    "prereq_e": ["prerequisitesubjectsforestudents", "estudents", "englishbasis", "e学生"],
    "reregister": ["reregistrationpossible", "reregistration", "再履修", "再登録"],
    "pf": ["pfevaluation", "passfail", "pf", "p/f"],
}


def match_header_row(values: Iterable[Any], aliases: dict[str, list[str]], required: set[str]) -> dict[str, int] | None:
    normalized = [key_text(v) for v in values]
    mapping: dict[str, int] = {}
    for field, candidates in aliases.items():
        for idx, cell in enumerate(normalized):
            if not cell:
                continue
            if any(candidate in cell for candidate in candidates):
                mapping[field] = idx
                break
    return mapping if required.issubset(mapping) else None


def find_header(ws, aliases: dict[str, list[str]], required: set[str], scan_rows: int = 60):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_rows), values_only=True), start=1):
        mapping = match_header_row(row, aliases, required)
        if mapping:
            return row_idx, mapping
    return None, None


def parse_timetable(path: Path, college: str) -> dict[str, dict[str, Any]]:
    # read_only=False so cell hyperlinks (direct syllabus URLs) stay reachable.
    wb = load_workbook(path, read_only=False, data_only=True)
    sections: dict[str, dict[str, Any]] = {}

    try:
        for ws in wb.worksheets:
            header_row, columns = find_header(ws, HEADER_ALIASES, {"class_code", "subject_cd", "name"})
            if not header_row or not columns:
                continue

            carry: dict[str, Any] = {}
            carry_fields = {"term", "class_code", "subject_cd", "name", "instructor", "language", "semester", "field", "area", "mode"}
            for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                if not any(clean_text(cell.value) for cell in row):
                    carry.clear()
                    continue

                def raw_val(field: str):
                    idx = columns.get(field)
                    return row[idx].value if idx is not None and idx < len(row) else None

                def val(field: str):
                    current = raw_val(field)
                    if field in carry_fields:
                        if clean_text(current):
                            carry[field] = current
                        else:
                            current = carry.get(field)
                    return current

                class_code = code_text(val("class_code"))
                subject_cd = code_text(val("subject_cd"))
                name = clean_text(val("name"))
                if not class_code or not subject_cd or not name:
                    continue

                subject_code = base_subject_code(subject_cd)
                if not subject_code:
                    continue

                section = sections.setdefault(class_code, {
                    "college": college,
                    "classCode": class_code,
                    "subjectCd": subject_cd,
                    "subjectCode": subject_code,
                    "name": name,
                    "instructor": clean_text(val("instructor")),
                    "language": clean_text(val("language")),
                    "term": normalize_term(val("term")),
                    "mode": normalize_mode(val("mode")),
                    "field": clean_text(val("field")),
                    "area": clean_text(val("area")),
                    "availableFromSemester": parse_semester_min(val("semester")),
                    "credits": parse_number(val("credits")),
                    "meetings": [],
                })
                syllabus_url = syllabus_link_from_cells(row, class_code)
                if syllabus_url and not section.get("syllabusUrl"):
                    section["syllabusUrl"] = syllabus_url

                day = normalize_day(val("day"))
                period = parse_period(val("period"))
                if day and period:
                    meeting = {
                        "day": day,
                        "period": period,
                        "classroom": clean_text(val("classroom")),
                        "mode": normalize_mode(val("mode")),
                    }
                    if meeting not in section["meetings"]:
                        section["meetings"].append(meeting)

                # Some sheets repeat metadata only on the first row; keep first non-empty value.
                for key, current in [
                    ("instructor", clean_text(val("instructor"))),
                    ("language", clean_text(val("language"))),
                    ("field", clean_text(val("field"))),
                    ("area", clean_text(val("area"))),
                ]:
                    if not section[key] and current:
                        section[key] = current
                if section["availableFromSemester"] is None:
                    section["availableFromSemester"] = parse_semester_min(val("semester"))
                if section["credits"] is None:
                    section["credits"] = parse_number(val("credits"))

    finally:
        wb.close()

    if not sections:
        raise ValueError("No timetable rows were recognized. The APU spreadsheet layout may have changed.")
    return sections


def parse_subject_list(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    subjects: dict[str, dict[str, Any]] = {}

    try:
        for ws in wb.worksheets:
            header_row, columns = find_header(ws, SUBJECT_HEADER_ALIASES, {"subject_code", "name", "credits"}, scan_rows=100)
            if not header_row or not columns:
                continue

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                def val(field: str):
                    idx = columns.get(field)
                    return row[idx] if idx is not None and idx < len(row) else None

                raw_code = code_text(val("subject_code"))
                code = base_subject_code(raw_code)
                name = clean_text(val("name"))
                credits = parse_number(val("credits"))
                if not code or not name or credits is None:
                    continue

                subjects[code] = {
                    "subjectCode": code,
                    "name": name,
                    "credits": credits,
                    "availableFromSemester": parse_semester_min(val("semester")),
                    "field": clean_text(val("field")),
                    "area": clean_text(val("area")),
                    "prerequisites": {
                        "JST": clean_text(val("prereq_jst")),
                        "JAT": clean_text(val("prereq_jat")),
                        "E": clean_text(val("prereq_e")),
                    },
                    "reregister": clean_text(val("reregister")),
                    "pfEvaluation": clean_text(val("pf")),
                }

    finally:
        wb.close()

    return subjects


def build_normalized(college: str, timetable_path: Path, subject_path: Path) -> dict[str, Any]:
    academic_term = detect_academic_term(timetable_path)
    sections = parse_timetable(timetable_path, college)
    apply_syllabus_links(sections.values(), academic_term.get("year"))
    metadata = parse_subject_list(subject_path)

    subjects: dict[str, dict[str, Any]] = {}
    for section in sections.values():
        code = section["subjectCode"]
        meta = metadata.get(code, {})
        if meta:
            section["credits"] = meta.get("credits") or section.get("credits")
            section["availableFromSemester"] = meta.get("availableFromSemester") or section.get("availableFromSemester")
            section["field"] = meta.get("field") or section.get("field", "")
            section["area"] = meta.get("area") or section.get("area", "")
        section["creditsEstimated"] = section.get("credits") is None
        if section.get("credits") is None:
            section["credits"] = 2.0

        subject = subjects.setdefault(code, {
            "subjectCode": code,
            "name": meta.get("name") or section["name"],
            "credits": section["credits"],
            "creditsEstimated": section["creditsEstimated"],
            "availableFromSemester": section.get("availableFromSemester"),
            "field": section.get("field", ""),
            "area": section.get("area", ""),
            "prerequisites": meta.get("prerequisites", {}),
            "reregister": meta.get("reregister", ""),
            "pfEvaluation": meta.get("pfEvaluation", ""),
            "sections": [],
        })
        subject["sections"].append(section)
        if not subject.get("availableFromSemester") and section.get("availableFromSemester"):
            subject["availableFromSemester"] = section["availableFromSemester"]

    subject_list = sorted(subjects.values(), key=lambda x: (x.get("field", ""), x["name"].lower()))
    section_list = sorted(sections.values(), key=lambda x: (x["name"].lower(), x["classCode"]))
    return {
        "schemaVersion": NORMALIZED_SCHEMA_VERSION,
        "college": college,
        "curriculum": 2023,
        "term": academic_term["label"],
        "academicYear": academic_term["year"],
        "academicSeason": academic_term["season"],
        "academicTermDetected": academic_term["detected"],
        "sourceVersion": DATA_VERSION if academic_term["label"] == "AY2026 Fall" else academic_term["label"],
        "officialUrls": OFFICIAL_FILES[college],
        "subjects": subject_list,
        "sections": section_list,
        "stats": {
            "subjects": len(subject_list),
            "sections": len(section_list),
            "subjectMetadataMatched": sum(1 for s in subject_list if not s["creditsEstimated"]),
        },
    }


def source_paths(college: str) -> tuple[Path, Path]:
    prefix = college.lower()
    return SOURCE_DIR / f"{prefix}_timetable.xlsx", SOURCE_DIR / f"{prefix}_subjects.xlsx"


def normalized_path(college: str) -> Path:
    return NORMALIZED_DIR / f"{college.upper()}.json"


def download_file(url: str, destination: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 APU-Schedule-Builder/0.9"})
    with urllib.request.urlopen(request, timeout=30) as response:
        data = response.read()
    if len(data) < 1000:
        raise ValueError("Downloaded file was unexpectedly small.")
    destination.write_bytes(data)


def load_or_build_data(college: str, allow_download: bool = True) -> dict[str, Any]:
    college = college.upper()
    if college not in OFFICIAL_FILES:
        raise ValueError("College must be APS, APM, or ST.")

    cached = normalized_path(college)
    timetable_path, subject_path = source_paths(college)
    if cached.exists():
        data = json.loads(cached.read_text(encoding="utf-8"))
        if data.get("schemaVersion") == NORMALIZED_SCHEMA_VERSION:
            # Syllabus URL mappings can be updated independently from timetable data.
            attach_syllabus_links(data)
            return data
        if timetable_path.exists() and subject_path.exists():
            data = build_normalized(college, timetable_path, subject_path)
            cached.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            return data
        # Old cache can still receive verified manual links even if the source XLSX is unavailable.
        attach_syllabus_links(data)
        return data

    if allow_download:
        if not timetable_path.exists():
            download_file(OFFICIAL_FILES[college]["timetable"], timetable_path)
        if not subject_path.exists():
            download_file(OFFICIAL_FILES[college]["subjects"], subject_path)

    if not timetable_path.exists() or not subject_path.exists():
        raise FileNotFoundError("Both timetable and subject-list XLSX files are required.")

    data = build_normalized(college, timetable_path, subject_path)
    cached.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def invalidate_college(college: str) -> None:
    path = normalized_path(college)
    if path.exists():
        path.unlink()


def section_slots(section: dict[str, Any]) -> set[tuple[str, str, int]]:
    term = section.get("term", "SEMESTER")
    quarters = ("Q1", "Q2") if term == "SEMESTER" else ((term,) if term in {"Q1", "Q2"} else ())
    slots: set[tuple[str, str, int]] = set()
    for meeting in section.get("meetings", []):
        day = meeting.get("day")
        period = meeting.get("period")
        if day not in DAYS or not isinstance(period, int):
            continue
        for q in quarters:
            slots.add((q, day, period))
    return slots


def conflict(a: dict[str, Any], b: dict[str, Any]) -> bool:
    return bool(section_slots(a) & section_slots(b))


def blocked_slots_from_config(config: dict[str, Any]) -> set[tuple[str, str, int]]:
    blocked: set[tuple[str, str, int]] = set()
    for value in config.get("blockedSlots", []):
        if not isinstance(value, str):
            continue
        parts = value.split(":")
        if len(parts) != 3:
            continue
        quarter, day, period_text = parts
        try:
            period = int(period_text)
        except ValueError:
            continue
        if quarter in {"Q1", "Q2"} and day in DAYS and 1 <= period <= 6:
            blocked.add((quarter, day, period))
    return blocked


def max_credits_for_semester(level: int, accelerated: bool = False) -> int:
    if accelerated and level >= 3:
        return 24
    if level <= 2:
        return 18
    if level <= 6:
        return 20
    return 24


@dataclass
class State:
    chosen: list[dict[str, Any]] = field(default_factory=list)
    occupied: set[tuple[str, str, int]] = field(default_factory=set)
    credits: float = 0.0
    prefer_count: int = 0
    estimated_credit_count: int = 0

    def add(self, section: dict[str, Any], preferred: bool = False) -> "State":
        return State(
            chosen=self.chosen + [section],
            occupied=self.occupied | section_slots(section),
            credits=self.credits + float(section.get("credits", 0)),
            prefer_count=self.prefer_count + (1 if preferred else 0),
            estimated_credit_count=self.estimated_credit_count + (1 if section.get("creditsEstimated") else 0),
        )


def schedule_metrics(state: State) -> dict[str, Any]:
    days_by_quarter: dict[str, set[str]] = {"Q1": set(), "Q2": set()}
    periods: dict[tuple[str, str], list[int]] = {}
    for q, day, period in state.occupied:
        days_by_quarter[q].add(day)
        periods.setdefault((q, day), []).append(period)

    campus_days = max((len(v) for v in days_by_quarter.values()), default=0)
    gaps = 0
    max_gap = 0
    earliest = 6
    latest = 1
    for plist in periods.values():
        plist = sorted(set(plist))
        if plist:
            earliest = min(earliest, plist[0])
            latest = max(latest, plist[-1])
            day_gaps = [max(0, b - a - 1) for a, b in zip(plist, plist[1:])]
            gaps += sum(day_gaps)
            max_gap = max(max_gap, max(day_gaps, default=0))
    return {
        "campusDays": campus_days,
        "gaps": gaps,
        "maxGap": max_gap,
        "earliest": earliest if state.occupied else None,
        "latest": latest if state.occupied else None,
        "daysByQuarter": {k: sorted(v, key=DAYS.index) for k, v in days_by_quarter.items()},
    }


def state_score(state: State, config: dict[str, Any], variant: str) -> float:
    target = float(config.get("targetCredits", 18))
    metrics = schedule_metrics(state)
    day_off = set(config.get("daysOff", []))
    earliest_pref = int(config.get("earliestPeriod", 1))
    latest_pref = int(config.get("latestPeriod", 6))
    max_days_pref = int(config.get("maxCampusDays", 5))
    max_gap_pref = int(config.get("maxGap", 5))
    preferred_languages = {str(v).upper() for v in config.get("preferredLanguages", []) if str(v).strip()}

    if variant == "fewest_days":
        weights = {"target": 14, "prefer": 45, "days": 30, "gaps": 6, "dayoff": 24, "early": 8, "late": 6, "gap_over": 10, "language": 3}
    elif variant == "course_priority":
        weights = {"target": 12, "prefer": 100, "days": 8, "gaps": 3, "dayoff": 12, "early": 5, "late": 4, "gap_over": 5, "language": 4}
    else:
        weights = {"target": 18, "prefer": 70, "days": 16, "gaps": 6, "dayoff": 22, "early": 8, "late": 6, "gap_over": 8, "language": 4}

    score = 500.0
    score -= abs(target - state.credits) * weights["target"]
    if math.isclose(target, state.credits, abs_tol=0.01):
        score += 90
    score += state.prefer_count * weights["prefer"]
    score -= metrics["campusDays"] * weights["days"]
    if metrics["campusDays"] <= max_days_pref:
        score += 30
    score -= metrics["gaps"] * weights["gaps"]

    occupied_days = set(metrics["daysByQuarter"]["Q1"]) | set(metrics["daysByQuarter"]["Q2"])
    score -= len(day_off & occupied_days) * weights["dayoff"]
    if metrics["earliest"] is not None and metrics["earliest"] < earliest_pref:
        score -= (earliest_pref - metrics["earliest"]) * weights["early"]
    if metrics["latest"] is not None and metrics["latest"] > latest_pref:
        score -= (metrics["latest"] - latest_pref) * weights["late"]
    if metrics["maxGap"] > max_gap_pref:
        score -= (metrics["maxGap"] - max_gap_pref) * weights["gap_over"]
    if preferred_languages:
        language_matches = 0
        for section in state.chosen:
            tokens = {token for token in re.split(r"[^A-Z]+", str(section.get("language", "")).upper()) if token}
            if preferred_languages & tokens:
                language_matches += 1
        score += language_matches * weights["language"]
    score -= state.estimated_credit_count * 1.5
    return score


def eligible_sections(subject: dict[str, Any], semester_level: int) -> list[dict[str, Any]]:
    minimum = subject.get("availableFromSemester")
    if minimum and semester_level < int(minimum):
        return []
    return [s for s in subject.get("sections", []) if s.get("term") in {"SEMESTER", "Q1", "Q2"}]


def validate_fixed(fixed_sections: list[dict[str, Any]], max_credits: float) -> tuple[State | None, str | None]:
    state = State()
    seen_subjects: set[str] = set()
    for section in fixed_sections:
        if section["subjectCode"] in seen_subjects:
            return None, f"Fixed classes contain multiple sections of {section['name']}."
        if state.occupied & section_slots(section):
            return None, f"Fixed class {section['name']} conflicts with another fixed class."
        state = state.add(section)
        seen_subjects.add(section["subjectCode"])
    if state.credits > max_credits:
        return None, f"Fixed classes already exceed the {max_credits:g}-credit limit."
    return state, None


def solve_variant(data: dict[str, Any], config: dict[str, Any], variant: str, beam_size: int = 220) -> tuple[list[tuple[float, State]], list[str]]:
    semester_level = int(config.get("semesterLevel", 5))
    accelerated = bool(config.get("accelerated", False))
    hard_max = max_credits_for_semester(semester_level, accelerated)
    requested_max = float(config.get("maxCredits", hard_max))
    max_credits = min(float(hard_max), requested_max)
    target = min(float(config.get("targetCredits", 18)), max_credits)
    statuses = {str(k): str(v).upper() for k, v in config.get("statuses", {}).items()}
    autofill = bool(config.get("autofill", False))
    blocked_slots = blocked_slots_from_config(config)

    section_by_code = {s["classCode"]: s for s in data["sections"]}
    fixed_codes = [code_text(v) for v in config.get("fixedClassCodes", []) if code_text(v)]
    missing_fixed = [c for c in fixed_codes if c not in section_by_code]
    if missing_fixed:
        return [], [f"Fixed class code not found: {', '.join(missing_fixed)}"]

    fixed_sections = [section_by_code[c] for c in fixed_codes]
    for section in fixed_sections:
        minimum = section.get("availableFromSemester")
        if minimum and semester_level < int(minimum):
            return [], [f"Fixed class {section['name']} requires semester {int(minimum)} or later."]
        if section_slots(section) & blocked_slots:
            return [], [f"Fixed class {section['name']} is in a disabled time slot."]
    initial, error = validate_fixed(fixed_sections, max_credits)
    if error:
        return [], [error]
    assert initial is not None

    fixed_subjects = {s["subjectCode"] for s in fixed_sections}
    subjects = []
    errors: list[str] = []

    for subject in data["subjects"]:
        code = subject["subjectCode"]
        status = statuses.get(code, "NEUTRAL")
        if status == "EXCLUDE" or code in fixed_subjects:
            continue
        sections = eligible_sections(subject, semester_level)
        if not sections:
            if status == "MUST":
                errors.append(f"{subject['name']} is marked MUST but has no eligible AY2026 Fall section for semester {semester_level}.")
            continue
        available_sections = [section for section in sections if not (section_slots(section) & blocked_slots)]
        if not available_sections:
            if status == "MUST":
                errors.append(f"{subject['name']} is MUST, but all eligible sections use disabled time slots.")
            continue
        if status in {"MUST", "PREFER"} or autofill:
            subjects.append((subject, status, available_sections))

    if errors:
        return [], errors

    # Required and preferred subjects are processed first; neutral autofill comes later.
    subjects.sort(key=lambda item: ({"MUST": 0, "PREFER": 1, "NEUTRAL": 2}.get(item[1], 2), item[0]["name"]))
    if autofill:
        # A bounded neutral pool keeps the local solver responsive without hiding selected courses.
        selected = [x for x in subjects if x[1] != "NEUTRAL"]
        neutral = [x for x in subjects if x[1] == "NEUTRAL"]
        neutral.sort(key=lambda x: (x[0].get("availableFromSemester") or 1, x[0]["name"]))
        subjects = selected + neutral[:120]

    beam: list[State] = [initial]
    for subject, status, sections in subjects:
        next_beam: list[State] = []
        must = status == "MUST"
        preferred = status == "PREFER"
        for state in beam:
            if not must:
                next_beam.append(state)
            for section in sections:
                credits = float(section.get("credits", subject.get("credits", 2)))
                if state.credits + credits > max_credits + 1e-9:
                    continue
                slots = section_slots(section)
                if state.occupied & slots:
                    continue
                next_beam.append(state.add(section, preferred=preferred))

        if not next_beam:
            if must:
                return [], [f"{subject['name']} is MUST, but every section conflicts with required/fixed classes or the credit limit."]
            continue

        next_beam.sort(key=lambda s: state_score(s, {**config, "targetCredits": target}, variant), reverse=True)
        # Deduplicate by selected class codes.
        seen: set[tuple[str, ...]] = set()
        beam = []
        for state in next_beam:
            signature = tuple(sorted(s["classCode"] for s in state.chosen))
            if signature in seen:
                continue
            seen.add(signature)
            beam.append(state)
            if len(beam) >= beam_size:
                break

    ranked = sorted(
        ((state_score(s, {**config, "targetCredits": target}, variant), s) for s in beam),
        key=lambda x: x[0],
        reverse=True,
    )
    return ranked[:30], []


def result_from_state(label: str, score: float, state: State, config: dict[str, Any]) -> dict[str, Any]:
    metrics = schedule_metrics(state)
    target = float(config.get("targetCredits", 18))
    days_off = set(config.get("daysOff", []))
    occupied_days = set(metrics["daysByQuarter"]["Q1"]) | set(metrics["daysByQuarter"]["Q2"])
    explanations = []
    if math.isclose(state.credits, target, abs_tol=0.01):
        explanations.append(f"Target {target:g} credits reached")
    else:
        explanations.append(f"{state.credits:g} credits selected (target {target:g})")
    if state.prefer_count:
        explanations.append(f"{state.prefer_count} preferred course(s) included")
    blocked_count = len(blocked_slots_from_config(config))
    if blocked_count:
        explanations.append(f"{blocked_count} blocked time slot(s) respected")
    free_requested = days_off - occupied_days
    if free_requested:
        explanations.append("Requested day(s) off kept: " + ", ".join(sorted(free_requested, key=DAYS.index)))
    explanations.append(f"Maximum {metrics['campusDays']} campus day(s) in a quarter")
    if metrics["gaps"] == 0:
        explanations.append("No timetable gaps")
    elif metrics["maxGap"] <= int(config.get("maxGap", 5)):
        explanations.append(f"Longest gap is {metrics['maxGap']} period(s)")
    if "earliestPeriod" in config and metrics["earliest"] is not None and metrics["earliest"] >= int(config["earliestPeriod"]):
        explanations.append(f"No class before period {int(config['earliestPeriod'])}")
    if "latestPeriod" in config and metrics["latest"] is not None and metrics["latest"] <= int(config["latestPeriod"]):
        explanations.append(f"No class after period {int(config['latestPeriod'])}")

    warnings = []
    if state.estimated_credit_count:
        warnings.append(f"{state.estimated_credit_count} course(s) use fallback 2-credit estimates because subject-list metadata did not match.")

    return {
        "label": label,
        "score": round(score, 1),
        "credits": state.credits,
        "courses": state.chosen,
        "metrics": metrics,
        "explanations": explanations,
        "warnings": warnings,
    }


def generate_schedules(data: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    variants = [
        ("BALANCED", "balanced"),
        ("FEWEST DAYS", "fewest_days"),
        ("COURSE PRIORITY", "course_priority"),
    ]
    pools: list[tuple[str, float, State]] = []
    all_errors: list[str] = []
    for label, variant in variants:
        ranked, errors = solve_variant(data, config, variant)
        all_errors.extend(errors)
        for score, state in ranked:
            pools.append((label, score, state))

    if not pools:
        return {"results": [], "errors": list(dict.fromkeys(all_errors)) or ["No valid schedule found."]}

    results = []
    seen: set[tuple[str, ...]] = set()
    # First try each variant's best unique schedule.
    for label, variant in variants:
        candidates = [(l, sc, st) for l, sc, st in pools if l == label]
        for _, score, state in candidates:
            signature = tuple(sorted(s["classCode"] for s in state.chosen))
            if signature in seen:
                continue
            seen.add(signature)
            results.append(result_from_state(label, score, state, config))
            break

    # Fill up to three with the best remaining unique alternatives.
    if len(results) < 3:
        for label, score, state in sorted(pools, key=lambda x: x[1], reverse=True):
            signature = tuple(sorted(s["classCode"] for s in state.chosen))
            if signature in seen:
                continue
            seen.add(signature)
            results.append(result_from_state("ALTERNATIVE", score, state, config))
            if len(results) >= 3:
                break

    return {"results": results[:3], "errors": []}
