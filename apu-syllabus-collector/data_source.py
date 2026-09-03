from __future__ import annotations

import json
import os
import re
import ssl
import subprocess
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

OFFICIAL_TIMETABLES = {
    "APS": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023APS_Curriculum_26Fall_260901.xlsx",
    "APM": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023APM_Curriculum_26Fall_260901.xlsx",
    "ST": "https://en.apu.ac.jp/academic/assets/file/class_info/timetable/2023ST_Curriculum_26Fall_260901.xlsx",
}
HEADER_ALIASES = {
    "class_code": ["coursecode", "classcode", "講義コード", "授業コード"],
    "name": ["subjectname", "coursename", "科目名"],
    "instructor": ["instructor", "teacher", "教員", "担当教員"],
    "term": ["term", "開講期間", "授業期間"],
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def key_text(value: Any) -> str:
    return re.sub(r"[\s\n\r\t._/()（）【】\[\]・:：-]+", "", clean(value).lower())


def code_text(value: Any) -> str:
    text = clean(value)
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def _is_cert_error(exc: BaseException) -> bool:
    pending = [exc]
    seen: set[int] = set()
    while pending:
        current = pending.pop()
        if id(current) in seen:
            continue
        seen.add(id(current))
        if isinstance(current, ssl.SSLCertVerificationError) or "certificate verify failed" in str(current).lower():
            return True
        for attr in ("reason", "__cause__", "__context__"):
            nested = getattr(current, attr, None)
            if isinstance(nested, BaseException):
                pending.append(nested)
    return False


def _windows_download(url: str, destination: Path) -> None:
    env = os.environ.copy()
    env["APU_COLLECTOR_URL"] = url
    env["APU_COLLECTOR_OUT"] = str(destination)
    script = "$ErrorActionPreference='Stop';$ProgressPreference='SilentlyContinue';Invoke-WebRequest -Uri $env:APU_COLLECTOR_URL -OutFile $env:APU_COLLECTOR_OUT -UseBasicParsing"
    subprocess.run(["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", script], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=90, env=env)


def download_file(url: str, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 APU-Syllabus-Collector/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=40) as response:
            data = response.read()
        if len(data) < 1000:
            raise ValueError("Downloaded timetable was unexpectedly small.")
        destination.write_bytes(data)
        return
    except Exception as exc:
        if os.name != "nt" or not _is_cert_error(exc):
            raise
    try:
        _windows_download(url, destination)
        if destination.stat().st_size < 1000:
            raise ValueError("Downloaded timetable was unexpectedly small.")
    except Exception:
        destination.unlink(missing_ok=True)
        raise


def detect_term(path: Path) -> tuple[int, str]:
    parts = [path.name]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            parts.append(ws.title)
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 35), values_only=True):
                parts.extend(clean(value) for value in row if clean(value))
    finally:
        wb.close()
    joined = " ".join(parts)
    match = re.search(r"(?:AY\s*)?(20\d{2})\s*(Spring|Fall)", joined, re.I)
    if match:
        return int(match.group(1)), match.group(2).title()
    match = re.search(r"(?<!\d)(\d{2})\s*(Spring|Fall)(?![A-Za-z])", joined, re.I)
    if match:
        return 2000 + int(match.group(1)), match.group(2).title()
    raise ValueError("Could not detect academic year / Spring-Fall from timetable.")


def _find_header(ws) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True), start=1):
        normalized = [key_text(value) for value in row]
        columns: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for index, cell in enumerate(normalized):
                if cell and any(alias in cell for alias in aliases):
                    columns[field] = index
                    break
        if {"class_code", "name"}.issubset(columns):
            return row_index, columns
    raise ValueError("Could not find Course code / Subject Name columns in timetable.")


def parse_timetable(path: Path, college: str) -> dict[str, Any]:
    year, season = detect_term(path)
    classes: dict[str, dict[str, str]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            try:
                header_row, columns = _find_header(ws)
            except ValueError:
                continue
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                def value(field: str) -> Any:
                    index = columns.get(field)
                    return row[index] if index is not None and index < len(row) else None
                code = code_text(value("class_code"))
                name = clean(value("name"))
                if not code or not name or not re.fullmatch(r"\d{4,6}", code):
                    continue
                classes.setdefault(code, {"classCode": code, "name": name, "instructor": clean(value("instructor")), "term": clean(value("term"))})
    finally:
        wb.close()
    if not classes:
        raise ValueError("No Class codes were recognized in timetable.")
    return {"college": college, "academicYear": year, "season": season, "term": f"AY{year} {season}", "classes": sorted(classes.values(), key=lambda item: (item["name"].lower(), item["classCode"]))}


def load_dataset(root: Path, college: str, refresh: bool = False) -> dict[str, Any]:
    college = college.upper()
    if college not in OFFICIAL_TIMETABLES:
        raise ValueError("College must be APS, APM, or ST.")
    source_dir = root / "data" / "source"
    source = source_dir / f"{college.lower()}_timetable.xlsx"
    cache = root / "data" / f"classes_{college.lower()}.json"
    if refresh:
        source.unlink(missing_ok=True)
        cache.unlink(missing_ok=True)
    if cache.exists() and source.exists():
        try:
            data = json.loads(cache.read_text(encoding="utf-8"))
            if isinstance(data, dict) and data.get("classes"):
                return data
        except (OSError, json.JSONDecodeError):
            pass
    if not source.exists():
        download_file(OFFICIAL_TIMETABLES[college], source)
    data = parse_timetable(source, college)
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return data
